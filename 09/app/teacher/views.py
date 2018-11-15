import shutil
import os
import zipfile
import json
from flask import flash, redirect, render_template, url_for, request,\
    current_app, send_from_directory, make_response, send_file
from datetime import datetime, timedelta
from flask_login import login_required, current_user
from ..upload_utils import secure_filename
from . import teacher
from ..auths import UserAuth
from .forms import *
from ..models.models import *
from flask_uploads import UploadNotAllowed
from openpyxl.utils.exceptions import InvalidFileException
from config import basedir
from openpyxl import Workbook
from sqlalchemy import not_


@teacher.before_request
@login_required
def before_request():
    pass


@teacher.route('/')
def index():
    courses = Teacher.query.filter_by(id=current_user.id).first().courses
    return render_template('teacher/index.html', courses=courses)


def download_file(directory, filename):
    response = make_response(send_from_directory(directory, filename, as_attachment=True))
    response.headers["Content-Disposition"] = "attachment; filename={}".format(filename.encode().decode('latin-1'))
    return response


@teacher.route('/<course_id>/course', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def set_course_info(course_id):
    # 教师课程信息
    form = CourseForm()
    course = Course.query.filter_by(id=course_id).first()
    if form.validate_on_submit():
        course.outline = form.outline.data
        course.teamsize_min = form.teamsize_min.data
        course.teamsize_max = form.teamsize_max.data
        course.no_miss = form.no_miss.data
        course.miss_1 = form.miss_1.data
        course.miss_2 = form.miss_2.data
        course.miss_3 = form.miss_3.data
        course.miss_4 = form.miss_4.data
        course.miss_5 = form.miss_5.data
        db.session.add(course)
        db.session.commit()

        if form.outlet_attachment.data:
            file = form.outlet_attachment.data
            filename = secure_filename(file.filename)
            filedir = os.path.join(basedir, 'uploads', str(course_id))
            # 总之有outlet就删了先。。
            for file_ in os.listdir(filedir):
                try:
                    name, _ = file_.split('.')
                    if name == 'outlet':
                        os.remove(os.path.join(filedir, file_))
                except ValueError:
                    pass

            file.save(os.path.join(filedir, filename))
            _, ext = filename.split('.')
            os.rename(os.path.join(filedir, filename), os.path.join(filedir, 'outlet.' + ext))

        flash('修改成功！', 'success')
        return redirect(url_for('teacher.set_course_info', course_id=course_id))
    form.outline.data = course.outline
    form.teamsize_min.data = course.teamsize_min
    form.teamsize_max.data = course.teamsize_max
    form.no_miss.data = course.no_miss
    form.miss_1.data = course.miss_1
    form.miss_2.data = course.miss_2
    form.miss_3.data = course.miss_3
    form.miss_4.data = course.miss_4
    form.miss_5.data = course.miss_5
    outlet_attachment = None
    filedir = os.path.join(basedir, 'uploads', str(course_id))
    for file in os.listdir(filedir):
        try:
            name, _ = file.split('.')
            if name == 'outlet':
                outlet_attachment = file
                break
        except ValueError:
            pass
    if request.args.get('action') == 'download':
        return send_from_directory(filedir, outlet_attachment, as_attachment=True)
    return render_template('teacher/course.html', course_id=course_id, form=form, course=course, nav='set_course_info', outlet_attachment=outlet_attachment)


@teacher.route('/<course_id>/resource', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def manage_resource(course_id):
    # 教师查看修改课程资源
    path = request.args.get('path')
    if not path:
        return redirect(url_for('teacher.manage_resource', course_id=course_id, path='/'))

    expand_path = os.path.join(current_app.config['UPLOADED_FILES_DEST'], 'resource', course_id, path[1:])
    if not os.path.exists(expand_path):
        # 没有文件夹？赶紧新建一个，真鸡儿丢人
        os.mkdir(expand_path)

    if request.method == 'POST' and 'file' in request.files:
        # 上传文件
        file = request.files['file']
        filename = secure_filename(file.filename)
        filepath = os.path.join(current_app.config['UPLOADED_FILES_DEST'], 'resource', course_id, path[1:], filename)
        if os.path.exists(filepath):
            flash('已经存在了同名文件', 'danger')
            return redirect(url_for('teacher.manage_resource', course_id=course_id, path=path))
        file.save(filepath)
        flash('上传成功！', 'success')
        return redirect(url_for('teacher.manage_resource', course_id=course_id, path=path))

    if request.method == 'POST' and 'dirname' in request.form:
        # 新建文件夹
        dirname = request.form['dirname']
        dirpath = os.path.join(current_app.config['UPLOADED_FILES_DEST'], 'resource', course_id, path[1:], dirname)
        if os.path.exists(dirpath):
            flash('已经存在了同名文件夹', 'danger')
            return redirect(url_for('teacher.manage_resource', course_id=course_id, path=path))
        os.mkdir(dirpath)
        flash('新建文件夹成功！', 'success')
        return redirect(url_for('teacher.manage_resource', course_id=course_id, path=path))

    if 'action' in request.form:
        if request.form.get('action') == 'delete':
            # 删除
            filepath = os.path.join(
                current_app.config['UPLOADED_FILES_DEST'],
                'resource',
                course_id,
                path[1:],
                request.form.get('filename')
            )
            if not os.path.exists(filepath):
                flash('文件不存在！', 'danger')
                return redirect(url_for('teacher.manage_resource', course_id=course_id, path=path))
            if os.path.isfile(filepath):
                os.remove(filepath)
            else:
                shutil.rmtree(filepath)
            flash('删除成功！', 'success')
            return redirect(url_for('teacher.manage_resource', course_id=course_id, path=path))

    if request.args.get('download'):
        # 下载
        filedir = os.path.join(
            current_app.config['UPLOADED_FILES_DEST'],
            'resource',
            course_id,
            path[1:])
        filename = request.args.get('filename')
        print(filename)
        if os.path.exists(os.path.join(filedir, filename)):
            return download_file(filedir, filename)
        else:
            flash('文件不存在！', 'danger')
            return redirect(url_for('teacher.manage_resource', course_id=course_id, path=path))

    files = []

    def sizeof_fmt(num, suffix='B'):
        for unit in ['', 'K', 'M', 'G', 'T', 'P', 'E', 'Z']:
            if abs(num) < 1024.0:
                return "%3.1f%s%s" % (num, unit, suffix)
            num /= 1024.0
        return "%.1f%s%s" % (num, 'Y', suffix)

    class file_attributes:
        name = ""
        size = ""
        create_time = datetime.min
        is_dir = False
        is_file = False

        def __init__(self, name, size, create_time, is_dir, is_file):
            self.name = name
            self.size = size
            self.create_time = create_time
            self.is_dir = is_dir
            self.is_file = is_file

    for file in os.scandir(expand_path):
        time = datetime.fromtimestamp(file.stat().st_mtime)
        files.append(file_attributes(file.name, sizeof_fmt(file.stat().st_size), time, file.is_dir(), file.is_file()))
    return render_template('teacher/resource.html', course_id=course_id, files=files, path=path, nav='manage_resource')


@teacher.route('/<course_id>/homework', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def homework(course_id):

    form = HomeworkForm()

    if request.args.get('get_teamhomework_all'):
        return get_teamhomework_all(course_id)
    elif request.args.get('get_score_all'):
        return get_score_all(course_id)

    if form.validate_on_submit():
        if form.weight.data > 100 or form.weight.data <= 0:
            flash('无效的权重', 'danger')
            return redirect(url_for('teacher.homework', course_id=course_id))
        begin_time, end_time = form.time.data.split(' - ')
        begin_time = datetime.strptime(begin_time, '%m/%d/%Y %H:%M')
        end_time = datetime.strptime(end_time, '%m/%d/%Y %H:%M')
        homework = Homework()

        homework.name = form.name.data
        homework.base_requirement = form.base_requirement.data
        homework.begin_time = begin_time
        homework.end_time = end_time
        homework.weight = form.weight.data
        homework.max_submit_attempts = form.max_submit_attempts.data
        homework.course_id = course_id

        db.session.add(homework)
        db.session.commit()

        flash('发布成功！', 'success')
        return redirect(url_for('teacher.homework', course_id=course_id))
    course = Course.query.filter_by(id=course_id).first()

    homework_list = Homework.query.filter_by(course_id=course_id).all()
    return render_template('teacher/homework.html', course_id=course_id, homeworks=homework_list, form=form, course=course, nav='homework')


# PudgeG负责：提交情况表导出↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓
@teacher.route('/<course_id>/homework/download', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def get_teamhomework_all(course_id):
    # 得到所有小队历次作业提交信息
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '团队累计提交作业情况'

    team_list = Team.query.filter_by(course_id=course_id).filter_by(status=2).all()
    # Team.team_list(course_id)
    homework_list = Homework.query.filter_by(course_id=course_id).all()
    # Homework.homework_list(course_id)

    # 总列数
    column_number = 2
    # 总行数
    row_number = 1

    # 第一行输入的内容

    worksheet.cell(row=1, column=1).value = '团队名称'
    worksheet.cell(row=1, column=2).value = '团队编号'

    for every_homework in homework_list:
        column_number += 1
        t = every_homework.order + 1
        worksheet.cell(row=1, column=column_number).value = '第' + str(t) + '次作业成绩'

    # 后续内容循环输入
    for every_team in team_list:
        row_number += 1
        worksheet.cell(row=row_number, column=1).value = every_team.team_name
        worksheet.cell(row=row_number, column=2).value = every_team.order + 1
        i = 2
        for every_homework in homework_list:
            i += 1
            _submission = Submission.query.filter_by(homework_id=every_homework.id).filter_by(team_id=every_team.id).all()
            if _submission:
                submission = Submission.query.filter_by(homework_id=every_homework.id).filter_by(team_id=every_team.id).all()[-1]
                worksheet.cell(row=row_number, column=i).value = submission.score
            else:
                worksheet.cell(row=row_number, column=i).value = '0'

    filename = 'all_homework_submit.xlsx'

    if not os.path.exists(os.path.join(basedir, 'homework')):
        os.mkdir(os.path.join(basedir, 'homework'))

    workbook.save(os.path.join(basedir, 'homework', filename))
    if os.path.isfile(os.path.join(basedir, 'homework', filename)):
        response = make_response(send_file(os.path.join(basedir, 'homework', filename)))
    else:
        flash('文件创建失败！', 'danger')
        return redirect(url_for('teacher.homework', course_id=course_id))
    response.headers["Content-Disposition"] = "attachment; filename=" + filename + ";"
    return response
# PudgeG负责：提交情况表导出↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑


# PudgeG负责：总成绩表导出↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓
@teacher.route('/<course_id>/plus/download', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def get_score_all(course_id):
    # 得到小队总成绩以及个人总成绩
    workbook = Workbook()

    worksheet_team = workbook.active
    worksheet_team.title = '小队总成绩'
    worksheet_team.append(['团队名称', '团队编号', '团队总成绩'])

    worksheet_member = workbook.create_sheet()
    worksheet_member.title = '成员成绩'
    worksheet_member.append(['团队名称', '团队编号', '学生姓名', '学生编号', '学生总成绩'])

    team_list = Team.query.filter_by(course_id=course_id).filter_by(status=2).all()
    # Team.team_list(course_id)
    homework_list = Homework.query.filter_by(course_id=course_id).all()
    # Homework.homework_list(course_id)

    # input_info = []
    for team in team_list:
        # Team表
        score = 0
        for every_homework in homework_list:
            _this_submission = Submission.query.filter_by(team_id=team.id).filter_by(homework_id=every_homework.id).first()
            if _this_submission is not None:
                this_submission = Submission.query.filter_by(team_id=team.id).filter_by(homework_id=every_homework.id).all()[-1]
                score += this_submission.score * every_homework.weight / 100

        plus_list = Plus.query.filter_by(course_id=course_id).all()
        for every_plus in plus_list:
            _this_plus = TeamPlus.query.filter_by(plus_id=every_plus.id).filter_by(team_id=team.id).first()
            if not _this_plus:
                score += _this_plus.score

        worksheet_team.append([team.team_name, team.order + 1, score])
        # input_info.append(submission_record)
    # worksheet_team.append(input_info)

    # input_info2 = []
    for team in team_list:
        # Member表
        member_list = TeamMember.query.filter_by(team_id=team.id).filter_by(status=1).all()

        # 团队负责人第一行
        # owner = Student.query.filter_by(id=team.owner_id).first()
        score = 0
        # 作业分数
        for every_homework in homework_list:
            _this_submission = Submission.query.filter_by(team_id=team.id).filter_by(homework_id=every_homework.id).first()
            if _this_submission is not None:
                this_submission = Submission.query.filter_by(team_id=team.id).filter_by(homework_id=every_homework.id).all()[-1]
                score += this_submission.score * every_homework.weight / 100

        # 加分项分数
        plus_list = Plus.query.filter_by(course_id=course_id).all()
        for every_plus in plus_list:
            _this_plus = TeamPlus.query.filter_by(plus_id=every_plus.id).filter_by(team_id=team.id).first()
            if not _this_plus:
                score += _this_plus.score

        # 签到分数
        attendance_list = Attendance.query.filter_by(course_id=course_id).all()
        times = 0
        total = 0
        for every_attendance in attendance_list:
            total += 1
            attendance_record = AttendanceStats.query.filter_by(attendance_id=every_attendance.id).filter_by(student_id=team.owner_id).first()
            if not attendance_record:
                times += 1
        score_attendance = get_attendance_score(times, total, course_id)

        worksheet_member.append([team.team_name, team.order + 1, team.owner.name, team.owner.id, score * team.owner_grade + score_attendance])
        # input_info2.append(submission_record)

        for every_member in member_list:
            # _every_member = Student.query.filter_by(id=every_member.student_id).first()
            # 签到分数
            attendance_list = Attendance.query.filter_by(course_id=course_id).all()
            times = 0
            total = 0
            for every_attendance in attendance_list:
                total += 1
                attendance_record = AttendanceStats.query.filter_by(attendance_id=every_attendance.id).filter_by(student_id=every_member.student_id).first()
                if not attendance_record:
                    times += 1
            score_attendance = get_attendance_score(times, total, course_id)

            worksheet_member.append([team.team_name, team.order + 1, every_member.student.name, every_member.student.id, score * every_member.grade + score_attendance])
            # input_info2.append(submission_record)
    # worksheet_member.append(input_info2)

    filename = 'all_score_final.xlsx'

    if not os.path.exists(os.path.join(basedir, 'homework')):
        os.mkdir(os.path.join(basedir, 'homework'))

    workbook.save(os.path.join(basedir, 'homework', filename))
    if os.path.isfile(os.path.join(basedir, 'homework', filename)):
        response = make_response(send_file(os.path.join(basedir, 'homework', filename)))
    else:
        flash('文件创建失败！', 'danger')
        return redirect(url_for('teacher.homework', course_id=course_id))
    response.headers["Content-Disposition"] = "attachment; filename=" + filename + ";"
    return response
# PudgeG负责：总成绩表导出↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑


@teacher.route('/<int:course_id>/homework/<int:homework_id>/attachment/<int:team_id>/<filename>', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def download_attachment(course_id, homework_id, team_id, filename):
    file_dir = os.path.join(current_app.config['UPLOADED_FILES_DEST'],
                            str(course_id),
                            str(homework_id),
                            str(team_id))

    # 取最新的一次上传和上传时的附件

    attachment_previous = Team \
        .query \
        .filter_by(id=team_id) \
        .filter(Team.submissions.any(homework_id=homework_id)) \
        .first() \
        .submissions[-1] \
        .attachment[0]

    filename_upload = attachment_previous.file_name
    file_uuid = attachment_previous.guid

    name_temp, ext = os.path.splitext(filename)

    if not os.path.exists(os.path.join(file_dir, file_uuid + ext)):
        return redirect(url_for('teacher.homework_detail',
                                course_id=course_id,
                                homework_id=homework_id))
    # 寻找保存目录下的uuid文件
    for i in os.listdir(file_dir):
        if i.startswith(str(file_uuid)):
            os.rename(os.path.join(file_dir, i), os.path.join(file_dir, filename_upload))

    return download_file(file_dir, filename_upload)


@teacher.route('/<int:course_id>/homework/<int:homework_id>', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def homework_detail(course_id, homework_id):
    # 教师查看作业详情&团队提交状况
    form = HomeworkForm()
    course = Course.query.filter_by(id=course_id).first()
    homework = Homework.query.filter_by(id=homework_id).first()

    # 取每个组最新的提交记录
    teams = Team.query.filter_by(course_id=course_id).all()

    submission_latest = {}

    for i in teams:
        if Submission.query.filter_by(homework_id=homework_id).filter_by(team_id=i.id).all():
            submission_latest[i.id] = Submission.query.filter_by(homework_id=homework_id).filter_by(team_id=i.id)[-1]

    if form.validate_on_submit():
        # 修改作业
        begin_time, end_time = form.time.data.split(' - ')
        begin_time = datetime.strptime(begin_time, '%m/%d/%Y %H:%M')
        end_time = datetime.strptime(end_time, '%m/%d/%Y %H:%M')

        homework.name = form.name.data
        homework.base_requirement = form.base_requirement.data
        homework.begin_time = begin_time
        homework.end_time = end_time
        homework.weight = form.weight.data
        homework.max_submit_attempts = form.max_submit_attempts.data
        homework.course_id = course_id

        flash('修改成功！', 'success')
        return redirect(url_for('teacher.homework_detail',
                                course_id=course_id,
                                homework_id=homework_id))

    # 批量下载学生作业
    if request.args.get('action') == 'multi_download':

        file_path = os.path.join(basedir, 'uploads', str(course_id), str(homework_id))

        if not os.path.exists(os.path.join(basedir, 'temp')):
            os.mkdir(os.path.join(basedir, 'temp'))
        save_path = os.path.join(basedir, 'temp', 'download.zip')

        submission_all = Submission \
            .query \
            .filter_by(homework_id=homework_id) \
            .order_by(Submission.id.desc()) \
            .all()

        # 建立 uuid与 上传时file_name的 键值对关系{uuid: file_name}
        rename_list = {}
        for submission_temp in submission_all:
            attachment_temp = Attachment.query.filter_by(submission_id=submission_temp.id).first()
            if attachment_temp:
                rename_list[str(attachment_temp.guid)] = str(attachment_temp.file_name)

        # 重命名文件并提供下载
        rename(file_path, rename_list)
        make_zip(file_path, save_path)
        #return send_from_directory(directory='/'.join(save_path.split('/')[:-1]), filename='download.zip', as_attachment=True)
        return download_file(os.sep.join(save_path.split(os.sep)[:-1]), 'download.zip')


    # json {'team_id':{'score': score, 'comments': comments}}
    # 提交评价和评论
    if request.form.get('action') == 'submit':
        _list = json.loads(request.form.get('data'))
        for team_id in _list:
            submission_temp = Submission.query.filter_by(homework_id=homework_id, team_id=team_id).order_by(Submission.id.desc()).first()
            submission_temp.score = _list[team_id]['score']
            submission_temp.comments = _list[team_id]['comments']
            db.session.add(submission_temp)
        db.session.commit()
        flash('成功提交评论！', 'success')
        return redirect(url_for('teacher.homework_detail', course_id=course_id, homework_id=homework_id))

    if request.form.get('action') == 'multi_upload':
        # 教师批量上传批改后的作业
        try:
            file = request.files.get('file')
            name_temp, ext = os.path.splitext(file.filename)
            # 保存文件在basedir/uploads/<course_id>/<homework_id>/teacher_corrected.ext (通过特定名字找下载)
            up_corrected.save(file,
                              folder=os.path.join(basedir, 'uploads', str(course_id), str(homework_id)),
                              name='teacher_corrected' + ext)
        except InvalidFileException:
            flash(u'附件类型不正确，请使用zip或rar', 'danger')
            return redirect(url_for('teacher.homework_detail', course_id=course_id, homework_id=homework_id))
        except UploadNotAllowed:
            flash(u'附件上传不允许')
            return redirect(url_for('teacher.homework_detail', course_id=course_id, homework_id=homework_id))
        # 可能加入全体广播 向全部学生广播教师修改作业已上传
        flash('上传成功', 'success')
        return redirect(url_for('teacher.homework_detail', course_id=course_id, homework_id=homework_id))

    teacher_corrected = False
    corrected_file_dir = os.path.join(basedir, 'uploads', str(course_id), str(homework_id))
    corrected_file_path = os.path.join(corrected_file_dir, 'teacher_corrected.zip')
    if os.path.exists(corrected_file_path):
        teacher_corrected = True

    if request.args.get('action') == 'download_corrected':
        return download_file(corrected_file_dir, 'teacher_corrected.zip')

    form.name.data = homework.name
    form.base_requirement.data = homework.base_requirement
    form.time.data = '{} - {}'.format(homework.begin_time.strftime('%m/%d/%Y %H:%M'),
                                      homework.end_time.strftime('%m/%d/%Y %H:%M'))
    form.weight.data = homework.weight
    form.max_submit_attempts.data = homework.max_submit_attempts

    return render_template('teacher/homework_detail.html',
                           course_id=course_id,
                           course=course,
                           form=form,
                           homework=homework,
                           teams=teams,
                           teacher_corrected=teacher_corrected,
                           submission_latest=submission_latest,
                           nav='homework')


# PudgeG负责：得到本次作业报表↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓
@teacher.route('/<int:course_id>/homework/<int:homework_id>/download', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def get_homework_report(course_id, homework_id):
    # 得到本次作业报表

    this_homework = Homework.query.filter_by(id=homework_id).first()
    team_this_course = Team.query.filter_by(course_id=this_homework.course_id).filter_by(status=2).all()
    # Team.team_list(this_homework.course_id)

    # submission_list = Submission.query.filter_by(homework_id=homework_id).all()
    # if len(submission_list) == 0:
    #     flash('无提交记录，请先催交！', 'danger')
    #     return redirect(request.args.get('next') or url_for('main.set_homework'))

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = this_homework.name + ' 提交情况'
    worksheet.append(['团队名称', '团队ID', '本次作业是否提交', '本次作业分数'])
    # input_info = []
    for team in team_this_course:
        _finished = Submission.query.filter_by(homework_id=homework_id).filter_by(team_id=team.id).all()
        if _finished:
            # 有提交记录，拿最后一个
            finished = Submission.query.filter_by(homework_id=homework_id).filter_by(team_id=team.id).all()[-1]
            worksheet.append([team.team_name, team.order + 1, 'Yes', finished.score])
        else:
            # 无提交记录
            worksheet.append([team.team_name, team.order + 1, 'No', 0])
        # input_info.append(homework_record)

        # def convert_status(status):
        #     switcher = {
        #         0: '作业未批改',
        #         1: '作业已批改'
        #     }
        #     return switcher.get(status, '其他')

    # worksheet.append(input_info)

    filename = 'this_homework.xlsx'
    if not os.path.exists(os.path.join(basedir, 'homework')):
        os.mkdir(os.path.join(basedir, 'homework'))
    workbook.save(os.path.join(basedir, 'homework', filename))
    if os.path.isfile(os.path.join(basedir, 'homework', filename)):
        response = make_response(send_file(os.path.join(basedir, 'homework', filename)))
    else:
        flash('文件创建失败！', 'danger')
        return redirect(url_for('teacher/teacher_teammanagement'))
    response.headers["Content-Disposition"] = "attachment; filename=" + filename + ";"
    return response
# PudgeG负责：得到本次作业报表↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑


# 上传老师批改后的作业
def teacher_corrected(course_id, homework_id):
    form = UploadCorrected()
    if form.validate_on_submit():
        if form.up_corrected:
            try:
                name_temp, ext = os.path.splitext(form.up_corrected.data.filename)
                # 保存文件在basedir/uploads/<course_id>/<homework_id>/teacher_corrected.ext (通过特定名字找下载)
                up_corrected.save(form.up_corrected,
                                  folder=os.path.join(basedir, 'uploads', str(course_id), str(homework_id)),
                                  name=u'teacher_corrected' + ext)
            except InvalidFileException:
                flash(u'附件类型不正确，请使用zip或rar', 'danger')
                return redirect(request.args.get('next') or url_for('teacher.teacher_corrected'))
            except UploadNotAllowed:
                flash(u'附件上传不允许')
                return redirect(request.args.get('next') or url_for('teacher.teacher_corrected'))
            # 可能加入全体广播 向全部学生广播教师修改作业已上传
            flash('上传成功')
            return redirect(url_for('teacher.teacher_corrected', form=form))
    return render_template('teacher/upload_corrected.html', form=form)


def add_member(student_id, team_id):
    team_member = TeamMember()
    team_member.team_id = team_id
    team_member.student_id = student_id
    team_member.status = 1
    db.session.add(team_member)
    delete_list = TeamMember.query.filter_by(status=2).filter_by(student_id=student_id).all()
    for a in delete_list:
        db.session.delete(a)
    db.session.commit()


# 用于生成zip文件
def make_zip(source_dir, output_filename):
    zipf = zipfile.ZipFile(output_filename, 'w')
    pre_len = len(os.path.dirname(source_dir))
    for parent, dirnames, filenames in os.walk(source_dir):
        for filename in filenames:
            pathfile = os.path.join(parent, filename)
            # 相对路径
            arcname = pathfile[pre_len:].strip(os.path.sep)
            zipf.write(pathfile, arcname)
    zipf.close()


# 用于下载前对文件重命名
def rename(source_dir, rename_dic):
    for parent, dirnames, filenames in os.walk(source_dir):
        for filename in filenames:
            name, ext = os.path.splitext(filename)
            if name in rename_dic.keys():
                os.rename(os.path.join(parent, filename), os.path.join(parent, rename_dic[name]))


@teacher.route('/teacher/<course_id>/givegrade_teacher/<homework_id>', methods=['GET', 'POST'])
def givegrade_teacher(course_id, homework_id):

    # 显示学生已提交的作业(显示最新的提交记录)
    submission = Submission.query.filter_by(homework_id=homework_id).filter_by(submit_status=1).all()
    team = Team.query.filter_by(course_id=course_id).all()

    #取每个组最新的提交记录
    submission_list = []
    for team_temp in team:
        submission_latest = Submission \
            .query \
            .filter_by(team_id=team_temp.id,
                       homework_id=homework_id) \
            .order_by(Submission.id.desc()) \
            .first()
        submission_list.append(submission_latest)

    homework_list = []
    for i in submission_list:
        team = Team.query.filter_by(id=i.team_id).first()
        homework_list.append({'team_id': i.team_id, 'team_name': team.team_name, 'text_content': i.text_content,
                              'score': i.score, 'comments': i.comments})

    # json [{'team_id':team_id, 'score': score, 'comments': comments}]
    # 提交评价和评论
    if request.method == 'POST' and request.form.get('action') == 'submit':
        _list = json.loads(request.form.get('data'))
        for dic in _list:
            for submission_temp in submission_list:
                if submission_temp.team_id == dic['team_id']:
                    submission_temp.score = dic['score']
                    submission_temp.comments = dic['comments']
                    db.session.add(submission_temp)
        db.session.commit()
        return redirect(url_for('teacher.givegrade_teacher', course_id=course_id, homework_id=homework_id))

    # 单个下载学生作业
    if request.method == 'POST' and request.form.get('action') == 'download':

        team_id = request.form.get('team_id')
        file_dir = os.path.join(current_app.config['UPLOADED_FILES_DEST'],
                                str(course_id),
                                str(homework_id),
                                str(team_id))

        # 取最新的一次上传和上传时的附件
        submission_previous = Submission \
            .query \
            .filter_by(team_id=team_id,
                       homework_id=homework_id) \
            .order_by(Submission.id.desc()) \
            .first()

        attachment_previous = None
        if submission_previous:
            attachment_previous = Attachment.query.filter_by(submission_id=submission_previous.id).first()

        # 无附件
        if not attachment_previous:
            flash('该组没有上传作业附件')
            return redirect(url_for('teacher.givegrade_teacher', course_id=course_id, homework_id=homework_id))
        else:

            filename_upload = attachment_previous.file_name
            file_uuid = attachment_previous.guid

            # 寻找保存目录下的uuid文件
            for i in os.listdir(file_dir):
                if i.startswith(str(file_uuid)):
                    os.rename(i, filename_upload)
            return download_file(file_dir, filename_upload)

    # 批量下载学生作业
    if request.method == 'POST' and request.form.get('action') == 'multi_download':

        file_path = os.path.join(basedir, 'uploads', str(course_id), str(homework_id))
        save_path = os.path.join(basedir, 'temp', 'download.zip')

        submission_all = Submission \
            .query \
            .filter_by(homework_id=homework_id) \
            .order_by(Submission.id.desc()) \
            .all()

        # 建立 uuid与 上传时file_name的 键值对关系{uuid: file_name}
        rename_list = {}
        for submission_temp in submission_all:
            attachment_temp = Attachment.query.filter_by(submission_id=submission_temp.id).first()
            rename_list[str(attachment_temp.guid)] = str(attachment_temp.file_name)

        # 重命名文件并提供下载
        rename(file_path, rename_list)
        make_zip(file_path, save_path)
        return download_file(file_path, 'download.zip')
    return render_template('teacher/givegrade_teacher.html', homework_list=homework_list)


# 教师往期课程作业
@teacher.route('/download', methods=['GET', 'POST'])
def see_class_before():

    # 下载往期课程作业

    course_id = request.args.get('course_id')

    file_path = os.path.join(basedir, 'uploads', str(course_id))
    save_path = os.path.join(basedir, 'temp', 'download.zip')

    if not os.path.exists(os.path.join(basedir, 'temp')):
        os.mkdir(os.path.join(basedir, 'temp'))

    if os.path.exists(file_path):
        make_zip(file_path, save_path)
        return download_file(os.path.join(basedir, 'temp'), 'download.zip')
    else:
        flash('这个课程没有附件作业保存！', 'danger')
        return redirect(url_for('teacher.see_class_before'))


@teacher.route('/<course_id>/team', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def team_manage(course_id):
    # 教师管理团队
    course = Course.query.filter_by(id=course_id).first()
    teams = Team.query.filter_by(course_id=course_id).all()
    form = MoveForm()
    if request.form.get('action') == 'accept':
        team = Team.query.filter_by(id=request.form.get('team_id')).first()
        if team:
            team.status = 2
            db.session.add(team)
            db.session.commit()
            flash("通过成功", "success")
            return redirect(url_for('teacher.team_manage', course_id=course_id))
        else:
            flash("找不到此团队", "danger")
            return redirect(url_for('teacher.team_manage', course_id=course_id))
    elif request.form.get('action') == 'reject':
        team = Team.query.filter_by(id=request.form.get('team_id')).first()
        if team:
            team.status = 3
            team.reject_reason = request.form.get('reason')
            db.session.add(team)
            db.session.commit()
            flash("拒绝成功", "success")
            return redirect(url_for('teacher.team_manage', course_id=course_id))
        else:
            flash("找不到此团队", "danger")
            return redirect(url_for('teacher.team_manage', course_id=course_id))
    elif request.form.get('action') == 'get_team_report':
        return get_team_report(course_id)

    unteamed_group = list(course.students)
    for team in teams:
        unteamed_group.remove(team.owner)

    members = TeamMember.query.filter(not_(TeamMember.status == 2)).filter(TeamMember.team.has(course_id=course_id)).all()

    for member in members:
        unteamed_group.remove(member.student)

    pending_teams = Team\
        .query\
        .filter_by(course_id=course_id)\
        .filter(not_(Team.status == 2))\
        .all()

    form.pending_teams.choices = [(r.id, r.team_name) for r in pending_teams]

    if form.validate_on_submit():
        print(2333)
        add_member(form.student.data, form.pending_teams.data)
        flash('移动成功！', 'success')
        return redirect(url_for('teacher.team_manage', course_id=course_id))

    return render_template('teacher/team.html',
                           course_id=course_id,
                           course=course,
                           teams=teams,
                           unteamed_group=unteamed_group,
                           form=form,
                           nav='team_manage')


# PudgeG负责:团队报表导出↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓
@teacher.route('/<course_id>/team/download', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def get_team_report(course_id):
    down_list = Team.query.filter_by(course_id=course_id).filter_by(status=2).all()
    # Team.team_list(course_id)
    if down_list is None:
        flash('没有已接受团队，请等待申请并批准！', 'danger')
        return redirect(request.args.get('next') or url_for('main.teacher_teammanagement'))
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '已接受团队信息'
    worksheet.append(['队伍名称', '队伍编号', '成员姓名', '成员编号', '成员角色'])
    # i = 0 表示队伍数量
    # input_info = []
    for team in down_list:
        member_list = TeamMember.query.filter_by(team_id=team.id).all()
        worksheet.append([team.team_name, team.order + 1, team.owner.name, team.owner_id, '团队负责人'])
        # input_info.append(input_record)
        # num_of_member = len(member_list)+1 表示每支队伍人员数量
        # i += 1
        for member in member_list:
            worksheet.append([team.team_name, team.order + 1, member.student.name, member.student_id, '普通成员'])
            # input_info.append(input_record)

    # worksheet.append(input_info)

    filename = 'all_team.xlsx'
    if not os.path.exists(os.path.join(basedir, 'team_manage')):
        os.mkdir(os.path.join(basedir, 'team_manage'))

    workbook.save(os.path.join(basedir, 'team_manage', filename))
    if os.path.isfile(os.path.join(basedir, 'team_manage', filename)):
        response = make_response(send_file(os.path.join(basedir, 'team_manage', filename)))
    else:
        flash('文件创建失败！', 'danger')
        return redirect(url_for('teacher/teacher_teammanagement'))
    response.headers["Content-Disposition"] = "attachment; filename=" + filename + ";"
    return response
# PudgeG负责:团队报表输出↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑


@teacher.route('/<course_id>/attendance', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def manage_attendance(course_id):
    course = Course.query.filter_by(id=course_id).first()
    attendance_list = Attendance.query.filter_by(course_id=course_id).all()
    if not attendance_list:
        attendance_available = True
    else:
        attendance_available = attendance_list[-1].time_end < datetime.now()  # 上次签到未过期，不可以发布下一次
    form = AttendanceForm()
    if form.validate_on_submit():
        if not attendance_available:
            flash("上一次签到未截止", "danger")
            return redirect(url_for('teacher.manage_attendance', course_id=course_id, attendance_available=attendance_available))
        new_attendance = Attendance()
        new_attendance.course_id = course_id
        new_attendance.info = form.info.data
        new_attendance.time_begin = datetime.now()
        new_attendance.time_end = datetime.now() + timedelta(minutes=form.time_delta.data)
        db.session.add(new_attendance)
        db.session.commit()
        flash("发布成功", "success")
        return redirect(url_for('teacher.manage_attendance',
                                course_id=course_id,
                                attendance_available=attendance_available,
                                attendance_list=attendance_list,
                                form=form,
                                course=course))
    return render_template('teacher/manage_attendance.html',
                           course_id=course_id,
                           attendance_available=attendance_available,
                           attendance_list=attendance_list,
                           form=form,
                           course=course,
                           nav='manage_attendance')


@teacher.route('/<course_id>/plus', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def add_plus(course_id):
    form = PlusForm()
    pluses = Plus.query.filter_by(course_id=course_id).all()
    if form.validate_on_submit():
        plus = Plus()
        plus.course_id = course_id
        plus.name = form.name.data
        plus.weight = form.weight.data
        db.session.add(plus)
        db.session.commit()
        flash('成功添加加分项', 'success')
        return redirect(url_for('teacher.add_plus', course_id=course_id))
    course = Course.query.filter_by(id=course_id).first()
    return render_template('teacher/plus.html', course_id=course_id, course=course, form=form, pluses=pluses, nav='add_plus')


@teacher.route('/<int:course_id>/plus_manage/<int:plus_id>', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def plus_manage(course_id, plus_id):
    # 加入学生信息
    teams_origin = Team.query.filter_by(course_id=course_id, status=2).all()
    teams = []
    for team in teams_origin:
        score = TeamPlus.query.filter_by(plus_id=plus_id, team_id=team.id).first()

        teams.append({
            'id': team.id,
            'team_name': team.team_name,
            'score': score.score if score else 0
        })

    # [{team_id:team_id, team_score:team_score}]
    if request.method == 'POST':
        _list = json.loads(request.form.get('data'))

        for dic in _list:
            # team plus
            tp = TeamPlus()
            tp.plus_id = plus_id
            tp.team_id = dic['team_id']
            tp.score = dic['team_score']
            tp.course_id = course_id
            db.session.add(tp)
        db.session.commit()
        flash('提交成功！', 'success')
        return redirect(url_for('teacher.plus_manage', course_id=course_id, plus_id=plus_id))
    plus = Plus.query.filter_by(id=plus_id).first()
    course = Course.query.filter_by(id=course_id).first()
    return render_template('teacher/plus_manage.html',
                           teams=teams,
                           course_id=course_id,
                           course=course,
                           plus=plus,
                           nav='add_plus')


# PudgeG负责：签到情况表导出↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓
@teacher.route('/<course_id>/attendance/download', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def get_attendence_all(course_id):
    # 得到所有小队历次作业提交信息
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '签到整体情况表'

    # student_list = SCRelationship.query.filter_by(course_id=course_id).join(Student, SCRelationship.student_id == Student.id).all()
    course = Course.query.filter_by(id=course_id).first()
    student_list = course.students
    attendance_list = Attendance.query.filter_by(course_id=course_id).all()

    # 总列数
    column_number = 2
    # 总行数
    row_number = 1

    # 第一行输入的内容

    worksheet.cell(row=1, column=1).value = '学生姓名'
    worksheet.cell(row=1, column=2).value = '学生编号'

    for every_attendance in attendance_list:
        column_number += 1
        t = every_attendance.order + 1
        worksheet.cell(row=1, column=column_number).value = '第' + str(t) + '次签到'
    column_number += 1
    worksheet.cell(row=1, column=column_number).value = '签到成绩'

    # 后续内容循环输入
    for every_student in student_list:
        row_number += 1
        worksheet.cell(row=row_number, column=1).value = every_student.name
        worksheet.cell(row=row_number, column=2).value = every_student.id
        i = 2
        times = 0
        total = 0
        for every_attendance in attendance_list:
            i += 1
            total += 1
            attendance_record = AttendanceStats.query.filter_by(attendance_id=every_attendance.id).filter_by(student_id=every_student.id).first()
            if attendance_record:
                worksheet.cell(row=row_number, column=i).value = 'Yes'
            else:
                times += 1
                worksheet.cell(row=row_number, column=i).value = 'No'
        i += 1
        worksheet.cell(row=row_number, column=i).value = get_attendance_score(times, total, course_id)

    filename = 'attendance_all.xlsx'

    if not os.path.exists(os.path.join(basedir, 'homework')):
        os.mkdir(os.path.join(basedir, 'homework'))

    workbook.save(os.path.join(basedir, 'homework', filename))
    if os.path.isfile(os.path.join(basedir, 'homework', filename)):
        response = make_response(send_file(os.path.join(basedir, 'homework', filename)))
    else:
        flash('文件创建失败！', 'danger')
        return redirect(url_for('teacher.homework', course_id=course_id))
    response.headers["Content-Disposition"] = "attachment; filename=" + filename + ";"
    return response


def get_attendance_score(times, total, course_id):
    course = Course.query.filter_by(id=course_id).first()
    if total == 0:
        return 0
    if times == 0:
        return course.no_miss
    elif times == 1:
        return course.miss_1
    elif times == 2:
        return course.miss_2
    elif times == 3:
        return course.miss_3
    elif times == 4:
        return course.miss_4
    else:
        return course.miss_5

# PudgeG负责：签到情况表导出↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑
