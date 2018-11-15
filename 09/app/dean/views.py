import os
import openpyxl
import uuid
from flask import render_template, redirect, url_for, flash, request
from flask_login import login_required
from datetime import date
from flask_uploads import UploadNotAllowed
from openpyxl.utils.exceptions import InvalidFileException
from config import basedir
from .. import db
from . import dean
from ..models.models import Semester, Course, SCRelationship, Student, TCRelationship, Teacher
from ..auths import UserAuth
from .forms import AddSemesterForm, CourseForm, ups


@dean.before_request
@login_required
def before_request():
    pass


def read_file(file_path):
    workbook = openpyxl.load_workbook(filename=file_path)  # 打开xls文件

    sheet_student = workbook.get_sheet_by_name('学生信息')
    sheet_teacher = workbook.get_sheet_by_name('老师信息')  # 通过sheet名字访问sheet
    student_info = []
    teacher_info = []
    for i in range(2, sheet_student.max_row + 1):
        student_list = {'id': sheet_student.cell(row=i, column=1).value,
                        'name': sheet_student.cell(row=i, column=2).value,
                        'password': 666}  # 学生初始密码 666
        student_info.append(student_list)
    for i in range(2, sheet_teacher.max_row + 1):
        teacher_list = {'id': sheet_teacher.cell(row=i, column=1).value,
                        'name': sheet_teacher.cell(row=i, column=2).value,
                        'teacher_info': sheet_teacher.cell(row=i, column=3).value,
                        'password': 666}  # 老师初始密码 666
        teacher_info.append(teacher_list)
    return student_info, teacher_info


@dean.route('/semester', methods=['GET', 'POST'])
@UserAuth.dean
def manage_semester():
    form = AddSemesterForm()
    if form.validate_on_submit():
        begin_time, end_time = form.time.data.split('-')
        month, day, year = begin_time.split('/')
        begin_time = date(int(year), int(month), int(day))
        month, day, year = end_time.split('/')
        end_time = date(int(year), int(month), int(day))
        if Semester.query.filter_by(id=form.id.data).first():
            flash('添加了重复的学期…', 'danger')
            return redirect(url_for('dean.manage_semester'))
        db.session.add(Semester(id=form.id.data, base_info=form.base_info.data,
                                begin_time=begin_time, end_time=end_time))
        db.session.commit()
        flash('添加成功！', 'success')
        return redirect(url_for('dean.manage_semester'))
    if request.args.get('action'):
        semester_id = request.args.get('semester_id')
        courses = Course.query.filter_by(semester_id=int(semester_id)).all()
        for course in courses:
            if request.args.get('action') == 'start':
                course.status = 1
            else:
                course.status = 0
            db.session.add(course)
        db.session.commit()
        flash('操作成功！', 'success')
        return redirect(url_for('dean.manage_semester'))

    semester_list = Semester.query.all()
    return render_template('dean/semester.html', form=form, semesters=semester_list, nav='manage_semester')


@dean.route('/course', methods=['GET', 'POST'])
@UserAuth.dean
def manage_course():
    semester_list = Semester.query.all()
    form = CourseForm()
    form.semester.choices = [(a.id, str(a.id / 100) + '学年第' + str(a.id % 100) + '学期') for a in semester_list]
    if 'action' in request.args:
        if request.args['action'] == 'delete':
            _course = Course.query.filter_by(id=int(request.args['id'])).first()
            if not _course:
                flash('找不到该课程', 'danger')
                return redirect(url_for('dean.manage_course'))
            db.session.delete(_course)
            db.session.commit()
            flash('删除成功', 'success')
            return redirect(url_for('dean.manage_course'))
        elif request.args['action'] == 'end':
            _course = Course.query.filter_by(id=int(request.args['id'])).first()
            if not _course:
                flash('找不到该课程', 'danger')
                return redirect(url_for('dean.manage_course'))
            _course.status = False
            db.session.commit()
            flash('结束成功', 'success')
            return redirect(url_for('dean.manage_course'))

    if form.validate_on_submit():

        # course的基本信息
        course = Course()
        course.name = form.name.data
        course.course_info = form.course_info.data
        course.place = form.place.data
        course.credit = int(form.credit.data)
        course.semester_id = form.semester.data
        course.outline = '无'
        course.teamsize_min = 1
        course.teamsize_max = 5
        course.no_miss = 0
        course.miss_1 = 0
        course.miss_2 = 0
        course.miss_3 = 0
        course.miss_4 = 0
        course.miss_5 = 0
        course.status = True

        try:

            if not os.path.exists(os.path.join(basedir, 'uploads')):
                os.mkdir(os.path.join(basedir, 'uploads'))
            # 上传文件处理
            print(os.getcwd() + '/uploads')
            filename = ups.save(form.stuff_info.data, name=str(uuid.uuid4())+".xlsx")
            file_path = os.path.join(basedir, 'uploads', filename)
            student_info, teacher_info = read_file(file_path=file_path)
        except UploadNotAllowed:
            flash('附件上传不允许！', 'danger')
            return redirect(request.args.get('next') or url_for('dean.manage_course'))
        except InvalidFileException:
            flash('附件类型不正确，请使用 xlsx！', 'danger')
            return redirect(request.args.get('next') or url_for('dean.manage_course'))

        db.session.add(course)
        db.session.commit()

        # 添加学生
        for i in student_info:
            student = Student.query.filter_by(id=i.get('id')).first()
            # 区分可修改和不可修改的数据
            if student is None:
                student = Student()
                student.password = str(i.get('password'))
                student.id = int(i.get('id'))
            student.name = i.get('name')
            db.session.add(student)

            # 添加学生课程关系
            course.students.append(student)
        # 添加教师
        for i in teacher_info:
            teacher = Teacher.query.filter_by(id=i.get('id')).first()
            # 区分可修改和不可修改的数据
            if teacher is None:
                teacher = Teacher()
                teacher.password = str(i.get('password'))
                teacher.id = int(i.get('id'))
            teacher.name = i.get('name')
            teacher.teacher_info = i.get('teacher_info')
            db.session.add(teacher)

            # 添加老师课程关系
            course.teachers.append(teacher)
        db.session.add(course)
        db.session.commit()
        os.remove(file_path)

        flash('添加成功！', 'success')
        return redirect(request.args.get('next') or url_for('dean.manage_course'))

    course_list = Course.query.all()  # 显示课程
    stuff_list = {}
    for course in course_list:
        student_list = course.students
        teacher_list = course.teachers
        stuff_list[course.id] = {
            'student_list': student_list,
            'teacher_list': teacher_list
        }

    return render_template('dean/course.html',
                           form=form,
                           courses=course_list,
                           semesters=semester_list,
                           stuff=stuff_list,
                           nav='manage_course')
