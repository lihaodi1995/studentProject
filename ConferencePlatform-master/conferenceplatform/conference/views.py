from django.shortcuts import render
from django.http import JsonResponse
from django.utils.datastructures import MultiValueDictKeyError
from django.db import transaction as database_transaction   
from django.db import IntegrityError, DatabaseError, models
from django.utils import timezone
from django.core.exceptions import *
import datetime
from openpyxl import Workbook
import json
from account.models import *
from .models import *
from .forms import *
from .utils import *
from account.decorators import user_has_permission


@user_has_permission('account.ConferenceRelated_Permission')
def add_conference(request):
    assert request.method == 'POST'
    form = ConferenceInfoForm(request.POST, request.FILES)
    if form.is_valid():
        with database_transaction.atomic():
            org = get_organization(request.user)
            assert org is not None
            try:
                subject = Subject.objects.get(name=form.cleaned_data['subject'])
            except Subject.DoesNotExist:
                return JsonResponse({'message': 'unknown subject'})

            conf = Conference.objects.create(
                organization=org, title=form.cleaned_data['title'], 
                subject=subject,
                template_no=form.cleaned_data['template_no'],
                introduction=form.cleaned_data['introduction'], 
                soliciting_requirement=form.cleaned_data['soliciting_requirement'],            
                register_requirement=form.cleaned_data['register_requirement'],                
                accept_due=form.cleaned_data['accept_due'],
                # modify_due=form.cleaned_data['modify_due'],
                register_start=form.cleaned_data['register_start'],
                register_due=form.cleaned_data['register_due'],
                conference_start=form.cleaned_data['conference_start'],
                conference_due=form.cleaned_data['conference_due'],
                #paper_template=form.cleaned_data['paper_template'],
                venue=form.cleaned_data['venue'],
            )
            conf.paper_template = request.FILES['paper_template']
            if 'modify_due' in request.POST:
                f = forms.DateTimeField()
                conf.modify_due = f.clean(request.POST['modify_due'])
            conf.save()
            if not valid_timepoints(conf):
                conf.delete()
                return JsonResponse({'message': 'timepoints not reasonable'})            

            activities_json_str = form.cleaned_data['activities']
            activities_json = json.loads(activities_json_str)
            
            try:
                for activity in activities_json:
                    add_activity(conf, activity)
            except KeyError as e:                
                raise e
            return JsonResponse({'message': 'success', 'data': {'pk': conf.pk}})
    else:
        return JsonResponse({'message': 'invalid uploaded data'})


@user_has_permission('account.NormalUser_Permission')
def paper_submit(request, id):
    assert request.method == 'POST'
    try:
        with database_transaction.atomic():
            conf = Conference.objects.get(pk=id)
            if conference_status(conf) != ConferenceStatus.accepting_submission:
                return JsonResponse({'message': 'wrong time range'})
            normal_user = request.user.normaluser
            s = Submission.objects.create(
                submitter=normal_user, institute=request.POST['institute'],                
                conference=conf,
                paper_name=request.POST['paper_name'], 
                paper_abstract=request.POST['paper_abstract'],
                authors=request.POST['authors'],
                state='S',
            )
            s.paper = request.FILES['paper']
            s.save()
            conf.num_submission = conf.num_submission + 1
            conf.save()
            return JsonResponse({'message': 'success', 'data': {'pk': s.pk}})
    except Conference.DoesNotExist:
        return JsonResponse({'message': 'invalid conference pk'})
    except MultiValueDictKeyError:
        return JsonResponse({'message': 'invalid uploaded data'})
    except IntegrityError:
        return JsonResponse({'message': 'multiple submission'})

@user_has_permission('account.NormalUser_Permission')
def conference_register(request, id):
    assert request.method == 'POST'
    try:
        with database_transaction.atomic():
            conf = Conference.objects.get(pk=id)
            if conference_status(conf) != ConferenceStatus.accepting_register:
                return JsonResponse({'message': 'wrong time range'})
            if request.POST['listen_only'] == 'false':
                paper_submission_id = int(request.POST['paper_id'])            
                paper_submission = Submission.objects.get(pk=paper_submission_id)
                if (paper_submission.submitter.pk != request.user.normaluser.pk 
                   or paper_submission.conference.pk != conf.pk):
                    return JsonResponse({'message': 'not matching'})
                if paper_submission.state != 'P':
                    return JsonResponse({'message': 'paper not passed'})
                r = RegisterInformation.objects.create(
                    user=request.user.normaluser,
                    conference=conf,
                    participants=request.POST['participants'],
                    submission=paper_submission,
                    # pay_voucher=request.FILES['pay_voucher'],
                )
            else:
                assert request.POST['listen_only'] == 'true'
                # listen only
                r = RegisterInformation.objects.create(
                    user=request.user.normaluser,
                    conference=conf,
                    participants=request.POST['participants'],
                )
            r.pay_voucher = request.FILES['pay_voucher']
            r.save()
            return JsonResponse({'message': 'success'})
    except Conference.DoesNotExist:
        return JsonResponse({'message': 'invalid conference pk'})
    except MultiValueDictKeyError:
        return JsonResponse({'message': 'invalid uploaded data'})
    except IntegrityError:
        return JsonResponse({'message': 'reduplicate register'})
    except Submission.DoesNotExist:
        return JsonResponse({'message': 'invalid paper id'})


@user_has_permission('account.ConferenceRelated_Permission')
def set_modify_due(request, id):
    assert request.method == 'POST'
    now = datetime.datetime.now()
    try:
        with database_transaction.atomic():
            conf = Conference.objects.get(pk=id)
            if get_organization(request.user).id != id:
                return JsonResponse({'message': 'permission error'})

            """ if conf.modify_due != None:
                return JsonResponse({'message': 'already set'}) """

            cleaner = forms.DateTimeField()
            due = cleaner.clean(request.POST['modify_due'])
            if due <= now:
                return JsonResponse({'message': 'too late'})
            if now >= conf.register_start:
                return JsonResponse({'message': 'review finished'})
            # 为了测试方便先注释掉：
            """ if now < conf.accept_due:
                return JsonResponse({'message': 'review not started'}) """
            
            conf.modify_due = due
            if not valid_timepoints(conf):
                conf.modify_due = None
                return JsonResponse({'message': 'timepoints not reasonable'})

            return JsonResponse({'message':'success'})
    except Conference.DoesNotExist:
        return JsonResponse({'message': 'invalid conference pk'})
    except MultiValueDictKeyError:
        return JsonResponse({'message': 'invalid uploaded data'})


def num_not_over(request):
    now = datetime.datetime.now()
    s = Conference.objects.filter(conference_due__gt=now)
    return JsonResponse({'message':'success', 'data': s.count()})

@user_has_permission('account.NormalUser_Permission')
def submit_after_modification(request, id):
    assert request.method == 'POST'
    ret = {'message': 'success', 'data': None}
    try:
        with database_transaction.atomic():
            prevsub = Submission.objects.get(pk=id)
            if request.user.pk != prevsub.submitter.user.pk:
                ret['message'] = 'not owned'
                return JsonResponse(ret)

            conf = prevsub.conference
            if conference_status(conf) != ConferenceStatus.reviewing_accepting_modification:
                ret['message'] = 'not accepting modification'
                return JsonResponse(ret)
            
            if prevsub.modified:
                ret['message'] = 'can\'t modify twice'
                return JsonResponse(ret)
            
            if prevsub.state != 'M':
                ret['message'] = 'need no modification'
                return JsonResponse(ret)
            
            textcleaner = forms.CharField(required=False)
            prevsub.modified_time = datetime.datetime.now()
            prevsub.modified_explain = textcleaner.clean(request.POST['explain'])
            prevsub.paper_old = prevsub.paper
            prevsub.paper_name_old = prevsub.paper_name
            prevsub.paper = request.FILES['paper']
            try:
                name = textcleaner.clean(request.POST['name'])
                if name != '':
                    prevsub.paper_name = name
            except MultiValueDictKeyError:
                pass

            try:
                abstract = textcleaner.clean(request.POST['abstract'])
                if abstract != '':
                    prevsub.paper_abstract = abstract
            except MultiValueDictKeyError:
                pass

            try:
                authors = textcleaner.clean(request.POST['authors'])
                if authors != '':
                    prevsub.authors = authors
            except MultiValueDictKeyError:
                pass

            prevsub.modified = True
            prevsub.save()

    except Submission.DoesNotExist:
        ret['message'] = 'invalid submission pk'
    except MultiValueDictKeyError:
        ret['message'] = 'invalid uploaded data'
    return JsonResponse(ret)

@user_has_permission('account.ConferenceRelated_Permission')
def export_submission_info(request, id):
    try:
        conf = Conference.objects.get(pk=id)
        if get_organization(request.user).pk != conf.organization.pk:
            return JsonResponse({'message': 'permission error'})

        sub_set = Submission.objects.filter(conference_id=id)
        wb = Workbook()
        ws = wb.active
        ws.append(['提交用户', '论文编号', '论文名称', '论文摘要', '提交时间', 
        '状态', '修改建议', '是否修改过', '修改时间', '修改说明', '第一作者', '第一作者单位',
        '第二作者', '第二作者单位', '第三作者', '第三作者单位', 
        '第四作者', '第四作者单位', '第五作者', '第五作者单位', 
        '通讯作者', '通讯作者单位'])
        for sub in sub_set:
            li = [sub.submitter.user.username, sub.pk ,sub.paper_name, 
                  sub.paper_abstract, sub.submit_time.strftime('%Y-%m-%d %H:%M:%S')]
            li.append(get_sheet_value_from_state(sub.state))
            li.append(sub.modification_advice)
            li.append('是' if sub.modified else '否')
            if sub.modified:
                li.extend([sub.modified_time.strftime('%Y-%m-%d %H:%M:%S'), sub.modified_explain])
            else:
                li.extend([None, None])
            print(sub.authors)
            authors = json.loads(sub.authors)
            for i in range(1, 6):
                pos = 'A' + str(i)
                if pos in authors:
                    li.extend([authors[pos]['name'], authors[pos]['institute']])
                else:
                    li.extend([None, None])
            if 'CA' in authors:
                li.extend([authors['CA']['name'], authors['CA']['institute']])
            else:
                li.extend([None, None])
            ws.append(li)
        fn = 'submission_info.xlsx'
        path = export_path(id, fn)
        url = export_url(id, fn)
        wb.save(path)
        return JsonResponse({'message': 'success', 'data': url})
    except Conference.DoesNotExist:
        return JsonResponse({'message': 'invalid conference pk'})
        
@user_has_permission('account.ConferenceRelated_Permission')
def export_register_info(request, id):
    """ try:
        conf = Conference.objects.get(pk=id)
        if get_organization(request.user).pk != conf.organization.pk:
            return JsonResponse({'message': 'permission error'})

        reg_set = RegisterInformation.objects.filter(conference_id=id)
        wb = Workbook()
        ws = wb.active
        ws.append(['提交用户', '聆听/论文编号', '参会者1', '性别', '是否住宿',
        '参会者2', '性别', '是否住宿', '参会者3', '性别', '是否住宿',
        '参会者4', '性别', '是否住宿', '参会者5', '性别', '是否住宿',
        '参会者6', '性别', '是否住宿',])
        for reg in reg_set:
            li = [reg.user.user.username,]
            if reg.submission == None:
                li.append('聆听')
            else:
                li.append(reg.submission.pk)
            participants = json.loads(reg.participants)
            if len(participants) > 6:
                return JsonResponse({'message': 'too much participants'})            
            for p in participants:
                li.extend([p['name'], p['gender'], '是' if p['reservation'] else '否'])
            ws.append(li)

        fn = 'register_info.xlsx'
        path = export_path(id, fn)
        url = export_url(id, fn)
        wb.save(path)
        return JsonResponse({"message": 'success', 'data': url})
    except Conference.DoesNotExist:
        return JsonResponse({'message': 'invalid conference pk'}) """
    try:
        conf = Conference.objects.get(pk=id)
        if get_organization(request.user).pk != conf.organization.pk:
            return JsonResponse({'message': 'permission error'})

        reg_set = RegisterInformation.objects.filter(conference_id=id)
        wb = Workbook()
        ws = wb.active
        ws.append(['姓名', '性别', '是否住宿', '提交用户', '聆听/论文编号'])
        for reg in reg_set:
            tail = [reg.user.user.username,]
            if reg.submission == None:
                tail.append('聆听')
            else:
                tail.append(reg.submission.pk)
            participants = json.loads(reg.participants)
            first = True
            for p in participants:
                li = [p['name'], p['gender'], '是' if p['reservation'] else '否']
                if first:
                    first = False
                    li.extend(tail)
                ws.append(li)
        fn = 'register_info.xlsx'
        path = export_path(id, fn)
        url = export_url(id, fn)
        wb.save(path)
        return JsonResponse({'message': 'success', 'data': url})
    except Conference.DoesNotExist:
        return JsonResponse({'message': 'invalid conference pk'})